from typing import List

import openpyxl
from PyQt6.QtWidgets import QMessageBox
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from entity import Question


def handle_excel_file(excel_file: str) -> dict:
    workbook: Workbook = openpyxl.load_workbook(filename=excel_file)
    sheetnames: List[str] = workbook.sheetnames
    return {sheetname: workbook[sheetname].max_row for sheetname in sheetnames}


def handle_row_question(excel_file: str, sheet_name: str, row_number: int) -> Question:
    workbook: Workbook = openpyxl.load_workbook(filename=excel_file)
    sheet: Worksheet = workbook[sheet_name]
    row = sheet[row_number]
    return Question.from_rows(sheet_name, sheet[1], row)[0]


def handle_save(sheet_name: str, row: int, index: int, content: str, excel_file: str):
    workbook: Workbook = openpyxl.load_workbook(filename=excel_file)
    sheet: Worksheet = workbook[sheet_name]
    sheet.cell(row, index + 1).value = content
    workbook.save(excel_file)


def warn_no_excel_file_chosen() -> None:
    message_box = QMessageBox()
    message_box.setIcon(QMessageBox.Icon.Critical)
    message_box.setWindowTitle("Error!")
    message_box.setText("No excel file chosen!")
    message_box.exec()
